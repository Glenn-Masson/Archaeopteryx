#GUI
import PySimpleGUI as sg   
import sys
import csv
import xlsxwriter
import string
import urllib
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.comments import Comment
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

### This makes the GUI and returns the entered values as values[key] e.g. values['outputroot'] 
layout = [
	[sg.Text('ArchaeopteryX', size = (50,1), font =('Helvetica',15), justification = 'left', background_color = 'white')],
	[sg.Text("A Really Clever & Helpful Add-on for Ensuring Outstanding Practice in HDX-MS",size = (65,1), font =('Helvetica',10), background_color = 'white')],
	[sg.Text('Built by Glenn Masson & Roger Williams', size = (50,1), font =('Helvetica',10), justification = 'left', background_color = 'white')],
	[sg.Text("ArchaeopteryX needs the Dynamx State Export file in an .csv format", background_color = 'white')],
    [sg.Text("Please direct to your DynamX State Export File (it's a .csv file):", background_color = 'white')],    
    [sg.InputText('DynamX_output.csv', key='inputfile', background_color = 'white'), sg.FileBrowse()],
    [sg.Text('This coluld be a unique I.D. to identify the experiment.', background_color = 'white')],              
    [sg.Text('Prefix Test:', size=(33, 1), background_color = 'white'), sg.InputText('prefix', key='outputroot')],      
    [sg.Text('Max Deuteration is the maximum fraction of Deuterium in your sample buffer:', background_color = 'white')],
    [sg.Text('Max Deuteration (in decimal format)', size=(33, 1), background_color = 'white'), sg.InputText('0.7', key='maxfracdeut')],      
    [sg.Text('Date of Analysis:', size=(33, 1), background_color = 'white'), sg.InputText('01.01.19', key='anal_date')],      
    [sg.Text('Date of Experiment:', size=(33, 1), background_color = 'white'), sg.InputText('25.12.18', key='ex_date')],      
    [sg.Text('Instrument:', size=(33, 1), background_color = 'white'), sg.InputText('Waters Synapt G2 Si', key='instrument')],         
    [sg.Text('Enter a significance threshold for differences between states A & B. For a 5% difference, enter 0.05', background_color = 'white')],
    [sg.Text('Threshold:', size=(33, 1), background_color = 'white'), sg.InputText('0.05', key='poshigh')],      
    [sg.Text('Number of Repeats:', size=(33, 1), background_color = 'white'), sg.InputText('3', key='number_repeats')],
    [sg.Text('These colour the spread sheet exchange values, default values are 0.1, 0.35, & 0.7.', background_color = 'white')],
    [sg.Text('Low Colour', size=(33, 1), background_color = 'white'), sg.InputText('0.1', key='low')],   
    [sg.Text('Mid Colour', size=(33, 1), background_color = 'white'), sg.InputText('0.35', key='mid')],      
    [sg.Text('High Colour', size=(33, 1), background_color = 'white'), sg.InputText('0.7', key='high')],     
    [sg.Submit(), sg.Cancel()],
    [sg.Text('version 0.1', size = (50,1), font =('Helvetica',10), justification = 'left', background_color = 'white')]
             ]      

window = sg.Window('ArchaeopteryX', default_element_size=(40, 1), background_color = 'white', grab_anywhere=False).Layout(layout)    
event, values = window.Read()   
window.Close()

###This converts the imported CSV to .xlsx 

def csvtoxlsx(csv_name, xlsx_name, directory, floats):
    """
    A function to convert a CSV file to XLSX so it can be used by openpyxl.
    csvname = file name of csv you want to convert (include .csv)
    xlsx_name = name you want to name the xlsx file (include .xlsx)
    cwd = directory to find csv file (can pass os.getcwd())
    floats = A list of column indexes in which floats appear
    """

    os.chdir(directory)

    f = open(csv_name, 'rt')
    csv.register_dialect('commas', delimiter=',')
    reader = csv.reader(f, dialect='commas')
    wb = Workbook()
    dest_filename = xlsx_name
    ws = wb.worksheets[0]
    ws.title = xlsx_name[:-5]

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):

            column_letter = get_column_letter((column_index + 1))

            if column_index in floats:
                s = cell
                #Handles heading row or non floats
                try:
                    s = float(s)
                    ws[('%s%s'%(column_letter, (row_index + 1)))].value = s

                except ValueError:
                    ws[('%s%s'%(column_letter, (row_index + 1)))].value = s

            elif column_index not in floats:
                #Handles openpyxl 'illigal chars'
                try:
                    ws[('%s%s'%(column_letter, (row_index + 1)))].value = cell

                except:
                    ws[('%s%s'%(column_letter, (row_index + 1)))].value = 'illigal char'



    wb.save(filename = dest_filename)

csv_name = values['inputfile']

csvtoxlsx(csv_name,'input.xlsx', os.path.abspath(os.curdir) ,[1,2,6,7,9,10,11,12,13,14,15])

###Now we take the converted csv file & the GUI inputs into the script #

inputfile = 'input.xlsx'
outputroot = values['outputroot']
maxfracdeut = float(values['maxfracdeut'])
date = values['anal_date']
expdate = values['ex_date']
instrument = values['instrument']
low = float(values['low'])
mid = float(values['mid'])
high = float(values['high'])
poshigh = float(values['poshigh'])
number_repeats = float(values['number_repeats'])

print
'Instrument', instrument
print
'poshigh', poshigh
al = Alignment(horizontal='center')
redFill = PatternFill(start_color='FF0000',
                      end_color='FF0000',
                      fill_type='solid')
cyanFill = PatternFill(start_color='00FFFF',
                       end_color='00FFFF',
                       fill_type='solid')
blueFill = PatternFill(start_color='0000FF',
                       end_color='0000FF',
                       fill_type='solid')
orangeFill = PatternFill(start_color='FFCC00',
                         end_color='FFCC00',
                         fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFF',
                        end_color='FFFFFF',
                        fill_type='solid')
wb = load_workbook(filename=inputfile, data_only=True)
ws = wb.active

# find out how many samples in the workbook
i = 1
column = 1
nsample = 0
lastsample = '!!!!!!'
samplelist = []
samplename = 'test'
firstlineofsamples = []
while i < 1000000:
    i = i + 1
    sample = ws.cell(row=i, column=column).value
    if sample is None:
        break
    if sample != lastsample:
        nsample = nsample + 1
        lastsample = sample
        firstlineofsamples.append(i)
        samplelist.append(sample)
lastline = i - 1
firstlineofsamples.append(i)
samplenumber = 0

for sample in samplelist:
    #    This section Labels our data, and creates the sheets.
    wbnew = load_workbook(filename=inputfile, data_only=True)
    wsold = wbnew.active
    wsold.title = 'DynamX_Data'
    wsnew = wbnew.create_sheet("Sumary", 0)
    wsnew.title = "Summary"
    wsnew['A2'] = 'Global Color scale'
    wsnew['A3'] = low
    wsnew['B3'] = mid
    wsnew['C3'] = high
    wsnew['E1'] = 'Date:'
    wsnew['E2'] = 'Experiment date:'
    wsnew['E3'] = 'Sample:'
    wsnew['E4'] = 'Instrument:'
    wsnew['C6'] = 'Max Frac Deut'
    wsnew['E6'] = maxfracdeut
    wsnew['E6'].alignment = al
    samplenumber = samplenumber + 1
    wsnew['H1'] = date
    wsnew['H1'].alignment = al
    wsnew['H2'] = expdate
    wsnew['H2'].alignment = al
    wsnew['H3'] = sample
    wsnew['H3'].alignment = al
    wsnew['H4'] = instrument
    wsnew['H4'].alignment = al
    wsnew['A7'] = 'Start'
    wsnew['A7'].alignment = al
    wsnew['B7'] = 'End'
    wsnew['B7'].alignment = al
    wsnew['C7'] = 'Mod'
    wsnew['C7'].alignment = al
    wsnew['D7'] = 'm/z'
    wsnew['D7'].alignment = al
    wsnew['E7'] = '#D'
    wsnew['E7'].alignment = al
    wsnew['F7'] = 'RT'
    wsnew['F7'].alignment = al
    wsnew['J1'] = 'SD'

    laststart = -10000
    lastend = -10000
    lastmod = -1000000
    lastprot = 'xxxxxxxxxx'
    i = firstlineofsamples[samplenumber - 1] - 1
    lastline = firstlineofsamples[samplenumber] - 1
    firstlineofpeps = []
    npeptides = 0
    begins = []
    ends = []
    mods = []
    errorpeptides = []
    while i < lastline:
        i = i + 1
        pepstart = ws.cell(row=i, column=2).value
        pepend = ws.cell(row=i, column=3).value
        modification = ws.cell(row=i, column=5).value
        if pepstart != laststart or pepend != lastend or modification != lastmod:
            laststart = pepstart
            lastend = pepend
            lastmod = modification
            firstlineofpeps.append(i)
            npeptides = npeptides + 1
            begins.append(pepstart)
            ends.append(pepend)
            mods.append(modification)
    firstlineofpeps.append(i + 1)
    i = 0
    state2beg = []
    useablepeptides = 0
    while i < npeptides:
        i = i + 1
        pepbegin = firstlineofpeps[i - 1]
        pepend = firstlineofpeps[i] - 1

        j = pepbegin - 1
        nstate = 0
        laststate = ws.cell(row=pepbegin, column=9).value
        while j < pepend:
            j = j + 1
            state = ws.cell(row=j, column=9).value
            if state != laststate:
                laststate = state
                nstate = nstate + 1
                state2beg.append(j)
    i = 0
    origcol = 10
    timecol = origcol
    times = []
    ntimes = 0
    while i < npeptides:
        i = i + 1
        pepbegin = firstlineofpeps[i - 1]
        pepend = firstlineofpeps[i] - 1
        j = pepbegin - 1
        while j < pepend:
            j = j + 1
            time = ws.cell(row=j, column=origcol).value
            k = 0
            new = True
            while k < ntimes:
                k = k + 1
                if time == times[k - 1]:
                    new = False
                    break
            if new:
                times.append(time)
                ntimes = ntimes + 1
    i = 0
    statecol = 9
    states = []
    nstates = 0
    while i < npeptides:
        i = i + 1
        pepbegin = firstlineofpeps[i - 1]
        pepend = firstlineofpeps[i] - 1
        j = pepbegin - 1
        while j < pepend:
            j = j + 1
            state = ws.cell(row=j, column=statecol).value
            k = 0
            new = True
            while k < nstates:
                k = k + 1
                if state == states[k - 1]:
                    new = False
                    break
            if new:
                states.append(state)
                nstates = nstates + 1

    usepeptide = []
    i = 0
    while i < npeptides:
        i = i + 1
        pepbegin = firstlineofpeps[i - 1]
        stateend = state2beg[i - 1] - 1
        ntimepts1 = stateend - pepbegin + 1
        if ntimepts1 != ntimes:
            print('**********************************************************')
            print('Error. reference state (', states[0], ') is lacking one or more time points ')
            print('Sample ', sample)
            print('Reference time points:', ntimepts1)
            print('State 2 time points:', ntimepts2)
            print('Peptide begin, end', begins[i - 1], ends[i - 1])
            for standard_time in times:
                timept = pepbegin - 1
                found = False
                while timept < stateend:
                    timept = timept + 1
                    time = ws.cell(row=timept, column=origcol).value
                    if standard_time == time:
                        found = True
                        break
                if not found:
                    print('Missing time point in reference:', standard_time)

            errorpep = (sample + " - " + " (" + str(begins[i - 1]) + ':' + str(ends[i - 1]) + ")")
            errorpeptides.append(errorpep)

            usepeptide.append(False)
        else:
            ntimepts2 = firstlineofpeps[i] - state2beg[i - 1]
            if ntimepts2 != ntimes:
                print('**********************************************************')
                print('Error. experimental state (', states[1], ') is lacking one or more time points ')
                print('Sample ', sample)
                print('Reference time points:', ntimepts1)
                print('State 2 time points:', ntimepts2)
                print('Peptide begin, end', begins[i - 1], ends[i - 1])
                for standard_time in times:
                    timept = state2beg[i - 1] - 1
                    stateend2 = firstlineofpeps[i] - 1
                    found = False
                    while timept < stateend2:
                        timept = timept + 1
                        time = ws.cell(row=timept, column=origcol).value
                        if standard_time == time:
                            found = True
                            break
                    if not found:
                        print('Missing time point in state 2:', standard_time)

                usepeptide.append(False)

                errorpep = (sample + " - " + " (" + str(begins[i - 1]) + ':' + str(ends[i - 1]) + ")")
                errorpeptides.append(errorpep)

            else:
                usepeptide.append(True)
                useablepeptides = useablepeptides + 1
    offset = ntimes - 1
    offset2 = 2 * offset + 2
    offset3 = offset2 + offset + 3
    offset4 = offset3 + offset
    aoffsetcol = 7
    wsnew.cell(row=6, column=aoffsetcol).value = 'a'
    wsnew.cell(row=6, column=aoffsetcol + offset + 1).value = 'b'
    state1 = states[0]
    state2 = states[1]
    wsnew.cell(row=6, column=aoffsetcol + 1).value = state1
    wsnew.cell(row=6, column=aoffsetcol + offset + 2).value = state2
    #    i=firstlineofpeps[0]-1
    #    i=-1 we want to skip time zero
    i = 0
    origcol = 10
    column = 6
    # write the column headings for time points
    #    while i < state2beg[0]-1:
    while i < ntimes - 1:
        i = i + 1
        column = column + 1
        # state a
        wsnew.cell(row=7, column=column).value = times[i]
        wsnew.cell(row=7, column=column).alignment = al
        # state b
        wsnew.cell(row=7, column=column + offset + 1).value = times[i]
        wsnew.cell(row=7, column=column + offset + 1).alignment = al
        # a - b
        wsnew.cell(row=7, column=column + offset2 + 1).value = times[i]
        wsnew.cell(row=7, column=column + offset2 + 1).alignment = al
        # SD of a
        wsnew.cell(row=7, column=column + offset3).value = times[i]
        wsnew.cell(row=7, column=column + offset3).alignment = al
        # SD of b
        wsnew.cell(row=7, column=column + offset4).value = times[i]
        wsnew.cell(row=7, column=column + offset4).alignment = al
    # write neaders for start end numbers for peptides in the differences block
    wsnew.cell(row=7, column=5 + offset2 + 1).value = 'Start'
    wsnew.cell(row=7, column=5 + offset2 + 1).alignment = al
    wsnew.cell(row=7, column=5 + offset2 + 2).value = 'End'
    wsnew.cell(row=7, column=5 + offset2 + 2).alignment = al
    # SDs
    wsnew.cell(row=7, column=6 + offset3 - 1).value = 'Start'
    wsnew.cell(row=7, column=6 + offset3 - 1).alignment = al
    wsnew.cell(row=7, column=6 + offset3).value = 'End'
    wsnew.cell(row=7, column=6 + offset3).alignment = al
    #    wsnew.cell(row=7,column=6+offset3).value='CS'
    #    wsnew.cell(row=7,column=6+offset3).alignment=al
    wsnew.cell(row=6, column=6 + offset3 + 1).value = 'a '
    wsnew.cell(row=5, column=6 + offset3 + 1).value = 'STDEVS'
    wsnew.cell(row=6, column=6 + offset3 + 2).value = state1
    wsnew.cell(row=6, column=6 + offset4 + 1).value = 'b '
    wsnew.cell(row=5, column=6 + offset4 + 1).value = 'STDEVS'
    wsnew.cell(row=6, column=6 + offset4 + 2).value = state2
    # write header for the differences block
    wsnew.cell(row=6, column=7 + offset2 + 1).value = 'a-b differences'

    outrow = 7
    uptakecol = 13
    sduptakecol = 14
    maxuptakecol = 7
    moverzcol = 8
    moverzoutcol = 4
    rtcol = 15

    # loop through peptides in the sample
    for peptide in range(0, npeptides):
        if usepeptide[peptide]:
            outrow = outrow + 1
            #            i=firstlineofpeps[peptide]-1 # need to skip first time point
            i = firstlineofpeps[peptide]
            outcol = 6
            avgrt = 0.0
            minrt = 99999
            maxrt = -9999
            nrt = 0
            #            ntimepoints=ntimes skip time zero
            ntimepoints = ntimes
            upperleftrow = 8
            upperleftcol = get_column_letter(7)

            lowerrightrow = npeptides + upperleftrow - 1
            lowerrightcol = get_column_letter(6 + offset)

            ulr_b = upperleftrow
            ulc_b = get_column_letter(6 + offset + 2)
            lrr_b = lowerrightrow
            lrc_b = get_column_letter(6 + offset + offset + 1)
            ulr_diff = upperleftrow
            ulc_diff = get_column_letter(6 + offset + offset + 4)
            lrr_diff = lowerrightrow
            lrc_diff = get_column_letter(6 + offset + offset + 2 + offset + 1)
            # write results for the a block
            uptakecolchar = get_column_letter(uptakecol)
            sduptakecolchar = get_column_letter(sduptakecol)
            while i < state2beg[peptide] - 1:
                i = i + 1
                rt = ws.cell(row=i, column=rtcol).value
                # get range of retention times for the observations of the peptide in the a and b states
                if rt > maxrt:
                    maxrt = rt
                if rt < minrt:
                    minrt = rt
                avgrt = avgrt + rt
                nrt = nrt + 1
                moverz = ws.cell(row=i, column=moverzcol).value
                #                moverzcolchar=get_column_letter(moverzcol)
                #                formulamoverz='='+wsold.title+'!'+moverzcolchar+str(i)+''
                wsnew.cell(row=outrow, column=moverzoutcol).value = moverz
                outcol = outcol + 1
                maxuptake = ws.cell(row=i, column=maxuptakecol).value
                mamaxuptakecolchar = get_column_letter(maxuptakecol)
                converttopercent = maxuptake / maxfracdeut
                # calculate percent deuteration
                # for a
                formuladeut = '=' + wsold.title + '!' + uptakecolchar + str(
                    i) + '/' + wsold.title + '!' + mamaxuptakecolchar + str(i) + '/' + '$E$6' + ''
                formulasddeut = '=' + wsold.title + '!' + sduptakecolchar + str(
                    i) + '/' + wsold.title + '!' + mamaxuptakecolchar + str(i) + '/' + '$E$6' + ''
                wsnew.cell(row=outrow, column=outcol).value = formuladeut
                wsnew.cell(row=outrow, column=outcol).number_format = '0%'
                wsnew.cell(row=outrow, column=outcol).alignment = al
                # sdv for a block
                wsnew.cell(row=outrow, column=outcol + offset3).value = formulasddeut
                wsnew.cell(row=outrow, column=outcol + offset3).number_format = '0%'
                wsnew.cell(row=outrow, column=outcol + offset3).alignment = al
                blockrange = upperleftcol + str(upperleftrow) + ':' + lowerrightcol + str(lowerrightrow)

                deltauptake = (ws.cell(row=i, column=uptakecol).value - ws.cell(row=i + ntimepoints,
                                                                                column=uptakecol).value) / converttopercent
                formuladeltauptake = '=(' + wsold.title + '!' + uptakecolchar + str(
                    i) + '-' + wsold.title + '!' + uptakecolchar + str(
                    i + ntimepoints) + ')/(' + wsold.title + '!' + mamaxuptakecolchar + str(i) + ')/' + '$E$6' + ''
                deltaoffset = offset2 + 1
                # write results for the a-b block
                wsnew.cell(row=outrow, column=outcol + deltaoffset).value = formuladeltauptake
                #               print 'formuladeltauptake',formuladeltauptake,'i',i,'outrow;',outrow,'outcol',outcol,'uptakecol',uptakecol
                #               sys.exit()
                wsnew.cell(row=outrow, column=outcol + deltaoffset).number_format = '0.0%'

            wsnew.cell(row=outrow, column=outcol + deltaoffset).alignment = al
            #            i=state2beg[peptide]-1# need to skip first time point
            i = state2beg[peptide]
            outcol = 6 + offset + 1
            # write resilts for the b block
            #### test
            #           while i < firstlineofpeps[peptide+1]-1:
            while i < firstlineofpeps[peptide + 1] - 1:
                i = i + 1
                rt = ws.cell(row=i, column=rtcol).value
                if rt > maxrt:
                    maxrt = rt
                if rt < minrt:
                    minrt - rt
                avgrt = avgrt + rt
                nrt = nrt + 1
                outcol = outcol + 1
                maxuptake = ws.cell(row=i, column=maxuptakecol).value
                converttopercent = maxuptake / maxfracdeut
                # calculate percent deutertation for the b bloc
                deut = ws.cell(row=i, column=uptakecol).value / converttopercent
                formuladeut = '=' + wsold.title + '!' + uptakecolchar + str(
                    i) + '/' + wsold.title + '!' + mamaxuptakecolchar + str(i) + '/' + '$E$6' + ''
                formulasddeut = '=' + wsold.title + '!' + sduptakecolchar + str(
                    i) + '/' + wsold.title + '!' + mamaxuptakecolchar + str(i) + '/' + '$E$6' + ''
                wsnew.cell(row=outrow, column=outcol).value = formuladeut
                wsnew.cell(row=outrow, column=outcol).alignment = al
                wsnew.cell(row=outrow, column=outcol).number_format = '0%'
                # sd
                wsnew.cell(row=outrow, column=outcol + offset3 - 1).value = formulasddeut
                wsnew.cell(row=outrow, column=outcol + offset3 - 1).number_format = '0%'
                wsnew.cell(row=outrow, column=outcol + offset3 - 1).alignment = al

            if nrt > 0:
                avgrt = avgrt / nrt
            else:
                print(
                    '@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
                print('error: nrt zero')
                print('i', i, 'firstlineofpeps[peptide+1]-1', firstlineofpeps[peptide + 1] - 1)
                print('peptide', peptide)
            minrtstr = "%5.2f" % minrt
            maxrtstr = "%-5.2f" % maxrt
            rtrange = minrtstr + '-' + maxrtstr
            # write out the range of retention times observed
            wsnew.cell(row=outrow, column=6).value = rtrange
            #### test
            start = begins[peptide]
            end = ends[peptide]
            mod = mods[peptide]
            # write start, end numbers for peptide
            wsnew.cell(row=outrow, column=1).value = start
            wsnew.cell(row=outrow, column=1).alignment = al
            wsnew.cell(row=outrow, column=5 + offset2 + 1).value = start
            wsnew.cell(row=outrow, column=5 + offset2 + 1).alignment = al
            wsnew.cell(row=outrow, column=5 + offset3).value = start
            wsnew.cell(row=outrow, column=5 + offset3).alignment = al
            

            wsnew.cell(row=outrow, column=2).value = end
            wsnew.cell(row=outrow, column=2).alignment = al
            wsnew.cell(row=outrow, column=3).value = mod
            #            wsnew.cell(row=outrow,column=3).alignment=al
            wsnew.cell(row=outrow, column=6 + offset2 + 1).value = end
            wsnew.cell(row=outrow, column=6 + offset2 + 1).alignment = al
            wsnew.cell(row=outrow, column=5 + offset3 + 1).value = end
            wsnew.cell(row=outrow, column=5 + offset3 + 1).alignment = al
            # write the maxdeuteration for the peptide
            wsnew.cell(row=outrow, column=5).value = maxuptake
            wsnew.cell(row=outrow, column=5).alignment = al
    # set conditional formatting for the new wrksheet
    blockrangecolorbar = 'A3:C3'
    wsnew.conditional_formatting.add(blockrangecolorbar,
                                     ColorScaleRule(start_type='num', start_value='$A$3', start_color='0000FF',
                                                    mid_type='num', mid_value='$B$3', mid_color='FFFF00',
                                                    end_type='num', end_value='$C$3', end_color='FF0000'))

    wsnew.conditional_formatting.add(blockrange,
                                     ColorScaleRule(start_type='num', start_value='$A$3', start_color='0000FF',
                                                    mid_type='num', mid_value='$B$3', mid_color='FFFF00',
                                                    end_type='num', end_value='$C$3', end_color='FF0000'))
    wsnew.conditional_formatting.add(blockrange,
                                     CellIsRule(operator='between', formula=['0', '0.2'], stopIfTrue=True,
                                                font=Font(color="FFFFFF")))
    wsnew.conditional_formatting.add(blockrange,
                                     CellIsRule(operator='between', formula=['0.2', '1.0'], stopIfTrue=True,
                                                font=Font(color="000000")))
    blockrangediff = ulc_diff + str(ulr_diff) + ':' + lrc_diff + str(lrr_diff)
    #
    blockrangeb = ulc_b + str(ulr_b) + ':' + lrc_b + str(lrr_b)
    wsnew.conditional_formatting.add(blockrangeb,
                                     ColorScaleRule(start_type='num', start_value='$A$3', start_color=''"0000FF",
                                                    mid_type='num', mid_value='$B$3', mid_color="FFFF00",
                                                    end_type='num', end_value='$C$3', end_color="FF0000"))
    wsnew.conditional_formatting.add(blockrangeb,
                                     CellIsRule(operator='between', formula=['0', '0.2'], stopIfTrue=True,
                                                font=Font(color="FFFFFF")))
    wsnew.conditional_formatting.add(blockrangeb,
                                     CellIsRule(operator='between', formula=['0.2', '1.0'], stopIfTrue=True,
                                                font=Font(color="000000")))
	
	
	
    poshigh = poshigh * 2.0
    posmiddle	=	poshigh / 2.0
    negmiddle = -1.0 * posmiddle
    neglow = -1.0 * poshigh

    intervals = [neglow, negmiddle, posmiddle, poshigh]
    i = 8 + offset + offset
    lastcol = i + 4
    j = -1
    keyvalues = []
    keycolors = [redFill, orangeFill, cyanFill, blueFill]
    while i < lastcol:
        i = i + 1
        j = j + 1
        colchar = get_column_letter(i)
        cellchar = colchar + '3'
        wsnew[cellchar] = intervals[j]
        wsnew[cellchar].fill = keycolors[j]
        keyvalues.append(cellchar)
    cellchar = keyvalues[3]
    formulaposhigh = '$' + cellchar[0] + '$' + cellchar[1]
    cellchar = keyvalues[2]
    formulaposmiddle = '$' + cellchar[0] + '$' + cellchar[1]
    cellchar = keyvalues[1]
    formulanegmiddle = '$' + cellchar[0] + '$' + cellchar[1]
    cellchar = keyvalues[0]
    formulaneglow = '$' + cellchar[0] + '$' + cellchar[1]
    lastcolchar = get_column_letter(lastcol)

    blockrangediff = ulc_diff + str(ulr_diff) + ':' + lrc_diff + str(lrr_diff)
    wsnew.conditional_formatting.add(blockrangediff,
                                     CellIsRule(operator='greaterThanOrEqual', formula=[formulaposhigh],
                                                stopIfTrue=True, fill=blueFill, font=Font(color="FFFFFF")))
    wsnew.conditional_formatting.add(blockrangediff,
                                     CellIsRule(operator='between', formula=[formulaposmiddle, formulaposhigh],
                                                stopIfTrue=True, fill=cyanFill,))
    wsnew.conditional_formatting.add(blockrangediff,
                                     CellIsRule(operator='between', formula=[formulanegmiddle, formulaposmiddle],
                                                stopIfTrue=True, fill=whiteFill))
    wsnew.conditional_formatting.add(blockrangediff,
                                     CellIsRule(operator='between', formula=[formulaneglow, formulanegmiddle],
                                                stopIfTrue=True, fill=orangeFill))
    wsnew.conditional_formatting.add(blockrangediff,
                                     CellIsRule(operator='lessThanOrEqual', formula=[formulaneglow], stopIfTrue=True,
                                                fill=redFill, font=Font(color="FFFFFF")))
    #	Calculate Average St Dev

    sdpccolstart = 12 + (3 * (ntimepoints - 1))
    sdpccolend = 11 + (5 * (ntimepoints - 1))
    sdpcrowsend = 7 + npeptides
    sdpccolstart_char = get_column_letter(sdpccolstart)
    sdpccolend_char = get_column_letter(sdpccolend)

    wsnew['K1'] = '=AVERAGE(' + sdpccolstart_char + '8:' + sdpccolend_char + str(sdpcrowsend) + ')'

    #	Calculate Mean difference

    difpccolstart = 10 + (2 * (ntimepoints - 1))
    difpccolend = 9 + (3 * (ntimepoints - 1))
    difpcrowend = 7 + npeptides
    difpccolstart_char = get_column_letter(difpccolstart)
    difpccolend_char = get_column_letter(difpccolend)
    wsnew['K2'] = '=AVERAGE(' + difpccolstart_char + '8:' + difpccolend_char + str(difpcrowend) + ')'

    #   Add the metadata front sheet
    metasheet = wbnew.create_sheet("Meta Data", 0)
    metacol = ws.column_dimensions['A']
    metacol.font = Font(bold=True)
    metarow = ws.row_dimensions[1]
    metarow.font = Font(size=12)
    metasheet['A1'] = 'HDX Meta Data Table'
    metasheet['A2'] = 'Data Set'
    metasheet['D2'] = sample
    metasheet['D2'].alignment = al
    metasheet['A3'] = 'Number of peptides'
    metasheet['D3'] = npeptides
    metasheet['D3'].alignment = al
    metasheet['A4'] = 'Average Standard Deviation'
    metasheet['D4'] = '=Summary!K1'
    metasheet['D4'].number_format = '0.00%'
    metasheet['D4'].alignment = al
    metasheet['A5'] = 'Timepoints (min)'
    
    # Converts the list 'times' into something readable by workbooks
    
    times.remove(0)
    times=[round(x, 3) for x in times]
    metasheet['D5'].value
    metasheet['D5'] = str(times)
    
    metasheet['A6'] = 'Significance Threshold'
    metasheet['D6'] = '+/-' + str((poshigh/2)*100) + '%'
    metasheet['D6'].alignment = al
    metasheet['A7'] = 'Mean Difference'
    metasheet['D7'] = '=Summary!K2'
    metasheet['D7'].number_format = '0.00%'
    metasheet['D7'].alignment = al
    metasheet['A8'] = 'Number of Repeats'
    metasheet['D8'] = number_repeats
    metasheet['D8'].alignment = al
    metasheet['A9'] = 'Number of Errors'
    metasheet['D9'] = len(errorpeptides)
    metasheet['D9'].alignment = al
    metasheet['A10'] = 'Peptides with Errors'
    for incorrectpeptide in errorpeptides:
        incorrectpeptide = incorrectpeptide + ' '
    metasheet['D10'] = ", ".join(errorpeptides)
   
    pymolsheet = wbnew.create_sheet("Pymol Generaor", 3)
   
    outfile = outputroot + '_' + sample + '.xlsx'
    wbnew.save(outfile)

sg.Popup('Success!', 'Result file name: ' + outfile)

